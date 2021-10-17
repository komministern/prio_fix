"""
Copyright (C) 2021 Oscar Franzén <oscarfranzen@protonmail.com>

This file is part of PRIO-fix Användarregistrering.

PRIO-fix Användarregistrering is free software: you can redistribute 
it and/or modify it under the terms of the GNU General Public License 
as published by the Free Software Foundation, either version 3 of the 
License, or (at your option) any later version.

PRIO-fix Användarregistrering is distributed in the hope that it will 
be useful, but WITHOUT ANY WARRANTY; without even the implied warranty 
of MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
GNU General Public License for more details.

You should have received a copy of the GNU General Public License
along with GCA Simulator.  If not, see <https://www.gnu.org/licenses/>.
"""

import logging
import sys
import os

import openpyxl
from openpyxl.styles import Font

from PySide6 import QtCore

logger = logging.getLogger(__name__)

class MyModel(QtCore.QObject):

    def __init__(self):
        super(MyModel, self).__init__()

        self.data = []

        # The following hack decides the current working directory. If the app is run as a bundled
        # PyInstaller executable, the cwd is given by the _MEIPASS attribute at runtime.
        if hasattr(sys, '_MEIPASS'):
            cwd = sys._MEIPASS
        else:
            cwd = os.path.abspath('.')
        resources = os.path.join(cwd, 'resources')
        self.template_file_path = os.path.join(resources, 'Mall Användarregistrering-1.98.xlsx')


    def quit(self):
        pass


    

    def collect_data_from_rows(self, path_to_file):

        wb = openpyxl.load_workbook(path_to_file)
        sheet = wb.active

        name_column = 8
        pnr_column = 9
        email_column = 25

        row_counter = 2

        name = sheet.cell(row=row_counter, column=name_column).value
        pnr = sheet.cell(row=row_counter, column=pnr_column).value
        email = sheet.cell(row=row_counter, column=email_column).value

        while name:

            self.data.append({'name': name, 'pnr': pnr, 'email': email})

            row_counter += 1

            name = sheet.cell(row=row_counter, column=name_column).value
            pnr = sheet.cell(row=row_counter, column=pnr_column).value
            email = sheet.cell(row=row_counter, column=email_column).value


    def process_row_data(self):

        for row_data in self.data:

            name = row_data['name']
            email = row_data['email']

            firstname, surname, possibleambiguity = self.extract_names(name, email)

            if firstname and surname:

                row_data['firstname'] = firstname
                row_data['surname'] = surname
                row_data['possibleambiguity'] = possibleambiguity

            else:

                print('Could not resolve "%s" with email address "%s".' % (name, email))


    def extract_names(self, name_string, email_string):

        name_string_lower = name_string.lower()

        email_string_pre_at = email_string.split('@')[0]
        email_components = email_string_pre_at.split('.')
        first_name_from_email_lower = email_components[0].lower()

        # Let's first check if that exact string is a substring of the name_string_lower

        if first_name_from_email_lower in name_string_lower:

            # Yes, most of our work is done!!!

            first_name_lower = first_name_from_email_lower
            surname_lower = name_string_lower.replace(first_name_lower, '').lstrip().rstrip()
            possible_ambiguity = False


        elif first_name_from_email_lower in name_string_lower.replace('å', 'a').replace('ä', 'a').replace('ö', 'o'):

            # The name contains å, ä or ö. These are not in the mail address for sure.

            index_of_name_first_character = name_string_lower.replace('å', 'a').replace('ä', 'a').replace('ö', 'o').index(first_name_from_email_lower)
            
            first_name_lower = name_string_lower[index_of_name_first_character:]
            surname_lower = name_string_lower.replace(first_name_lower, '').lstrip().rstrip()
            possible_ambiguity = False

        else:

            # The first name in the mail address is not in the name_string_lower.

            if len(name_string_lower.split()) == 2:

                # Simple name, the surname_lower should be the first one, and the name should be the second.

                first_name_lower = name_string_lower.split()[1]
                surname_lower = name_string_lower.split()[0]
                possible_ambiguity = True

            else:

                # We have simply not been able to extract the first and surname. Fuck it!

                first_name_lower = ''
                surname_lower = ''
                possible_ambiguity = True


        return self.capitalize(first_name_lower), self.capitalize(surname_lower), possible_ambiguity


    def create_registration_from_template(self, path_to_file):

        firstname_column = 1
        surname_columnn = 2
        pnr_column = 3
        email_column = 4
        prioname_column = 8


        boldFont = Font(bold=True)
        italicFont = Font(italic=True)
        bolditalicFont = Font(bold=True, italic=True, name='Arial', size='10')

        nbr_of_files = len(self.data) // 90 + 1
        data_counter = 0
        suspect_counter = 0

        for file_counter in range(nbr_of_files):
        
            wb = openpyxl.load_workbook(self.template_file_path)
            sheet = wb.active

            row_counter = 11

            while data_counter < len(self.data) and row_counter <= 100:

            # for row_dict in self.data:

                row_dict = self.data[data_counter]

                prioname = row_dict['name']
                firstname = row_dict['firstname']
                surname = row_dict['surname']
                pnr = row_dict['pnr']
                email = row_dict['email']
                suspected = row_dict['possibleambiguity']
                
                firstnamecellref = sheet.cell(row=row_counter, column=firstname_column)
                surnamecellref = sheet.cell(row=row_counter, column=surname_columnn)
                pnrcellref = sheet.cell(row=row_counter, column=pnr_column)
                emailcellref = sheet.cell(row=row_counter, column=email_column)
                prionamecellref = sheet.cell(row=row_counter, column=prioname_column)

                firstnamecellref.value = firstname
                surnamecellref.value = surname
                pnrcellref.value = pnr
                emailcellref.value = email.lower()

                # if firstname == '' and surname == '':
                #     prionamecellref.value = prioname

                if suspected:
                    prionamecellref.value = prioname

                    for cell in sheet[str(row_counter) + ':' + str(row_counter)]:
                        cell.font = bolditalicFont

                    suspect_counter += 1

                row_counter += 1
                data_counter += 1

            if file_counter > 0:
                modified_path_to_file = path_to_file[:-5] + '_' + str(file_counter + 1) + path_to_file[-5:]
                wb.save(modified_path_to_file)
            else:
                wb.save(path_to_file)

        return len(self.data), suspect_counter, nbr_of_files


    def capitalize(self, astring):
        if astring:

            s1 = '-'.join([s.capitalize() for s in astring.split('-')])
            s2 = ' '.join([s.capitalize() for s in s1.split(' ')])

            return s2

        else:

            return ''